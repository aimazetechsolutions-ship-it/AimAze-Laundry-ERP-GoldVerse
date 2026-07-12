from datetime import date, datetime, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from odoo import _, api, fields, models
from odoo.tools.misc import get_lang


class InteractiveAccountReport(models.AbstractModel):
    _inherit = "account.interactive.report"

    @api.model
    def _normalize_options(self, report_key, options=None):
        normalized = super()._normalize_options(report_key, options=options)
        normalized["account_ids"] = [
            int(account_id)
            for account_id in (options or {}).get("account_ids", [])
            if account_id
        ]
        return normalized

    @api.model
    def action_pdf(self, report_key, options=None):
        options = self._normalize_options(report_key, options)
        return self.env.ref("goldverse_premium_laundry_branding.action_report_interactive_account_pdf").report_action(
            self.env.company,
            data={"report_key": report_key, "options": options},
        )

    @api.model
    def goldverse_export_payload(self, report_key, options=None):
        options = self._normalize_options(report_key, options)
        payload = self.get_report(report_key, options)
        payload = dict(payload)
        payload["period_label"] = self._goldverse_period_label(options)
        payload["generated_at"] = fields.Datetime.context_timestamp(self, datetime.utcnow()).strftime("%d %b %Y, %I:%M %p")
        payload["filters"] = self._goldverse_export_filters(payload)
        payload["export_columns"] = self._goldverse_export_columns(payload)
        payload["export_lines"] = self._goldverse_export_lines(payload)
        return payload

    @api.model
    def goldverse_xlsx_content(self, report_key, options=None):
        payload = self.goldverse_export_payload(report_key, options)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = self._goldverse_safe_sheet_title(payload["title"])
        self._goldverse_build_xlsx(sheet, payload)
        output = BytesIO()
        workbook.save(output)
        return output.getvalue(), self._goldverse_export_filename(payload, "xlsx")

    @api.model
    def _goldverse_export_columns(self, payload):
        columns = []
        for column in payload.get("columns", []):
            columns.append(
                {
                    "key": column["key"],
                    "label": column.get("label") or (payload["title"] if column["key"] == "name" else ""),
                    "type": column.get("type") or "text",
                }
            )
        return columns

    @api.model
    def _goldverse_export_lines(self, payload):
        export_columns = payload["export_columns"]
        lines = []
        for line in payload.get("lines", []):
            level = line.get("level") or 1
            is_total = bool(line.get("is_total")) or line.get("type") in {"aged_total", "total", "aged_summary"}
            cells = []
            for column in export_columns:
                raw_value = (line.get("values") or {}).get(column["key"], "")
                is_number = column["type"] == "number"
                indent = max(level - 1, 0) * 14 if column["key"] in ("name", "move") else 0
                cells.append(
                    {
                        "key": column["key"],
                        "value": raw_value,
                        "display": self._goldverse_display_export_cell(raw_value, column, payload),
                        "number": is_number,
                        "indent": indent,
                    }
                )
            lines.append(
                {
                    "id": line.get("id"),
                    "type": line.get("type"),
                    "level": level,
                    "is_total": is_total,
                    "cells": cells,
                }
            )
        return lines

    @api.model
    def _goldverse_display_export_cell(self, value, column, payload):
        if column["type"] != "number":
            return value or ""
        amount = float(value or 0.0)
        return f"{amount:,.2f} {payload.get('currency_label') or ''}".strip()

    @api.model
    def _goldverse_period_label(self, options):
        date_from = options.get("date_from")
        date_to = options.get("date_to")
        try:
            start = fields.Date.from_string(date_from)
            end = fields.Date.from_string(date_to)
        except Exception:
            return f"{date_from or ''} to {date_to or ''}".strip()
        if options.get("period") == "month":
            return start.strftime("%B %Y")
        if options.get("period") == "quarter":
            return f"{start.strftime('%b')} - {end.strftime('%b %Y')}"
        if options.get("period") == "year":
            return str(start.year)
        return f"{start.strftime('%d %b %Y')} to {end.strftime('%d %b %Y')}"

    @api.model
    def _balance_sheet_statement(self, options):
        columns, lines = super()._balance_sheet_statement(options)
        retained_earnings, current_year_earnings = self._goldverse_balance_sheet_earnings(options)
        total_earnings = retained_earnings + current_year_earnings

        old_earnings = 0.0
        current_index = False
        for index, line in enumerate(lines):
            if line.get("id") == "current_year_earnings":
                old_earnings = (line.get("values") or {}).get("balance") or 0.0
                current_index = index
                break
        if current_index is False:
            return columns, lines

        replacement_lines = []
        if abs(retained_earnings) >= 0.005:
            replacement_lines.append(
                self._goldverse_earnings_line(
                    "retained_earnings",
                    _("Retained Earnings"),
                    retained_earnings,
                )
            )
        replacement_lines.append(
            self._goldverse_earnings_line(
                "current_year_earnings",
                _("Current Year Earnings"),
                current_year_earnings,
            )
        )
        lines[current_index : current_index + 1] = replacement_lines

        earnings_delta = total_earnings - old_earnings
        for line in lines:
            if line.get("id") in ("total_equity", "equity_total", "total_liabilities_equity"):
                values = line.get("values") or {}
                values["balance"] = (values.get("balance") or 0.0) + earnings_delta
        return columns, lines

    @api.model
    def _goldverse_balance_sheet_earnings(self, options):
        date_to = fields.Date.from_string(options.get("date_to") or fields.Date.context_today(self))
        if not isinstance(date_to, date):
            date_to = fields.Date.context_today(self)
        year_start = date(date_to.year, 1, 1)

        current_options = dict(options, date_from=fields.Date.to_string(year_start))
        retained_options = dict(options, date_from=False, date_to=fields.Date.to_string(year_start - timedelta(days=1)))

        retained = self._profit_and_loss_totals(retained_options)["net_profit"] if year_start > date.min else 0.0
        current = self._profit_and_loss_totals(current_options)["net_profit"]
        return retained, current

    @api.model
    def _goldverse_earnings_line(self, key, name, amount):
        return {
            "id": key,
            "name": name,
            "level": 2,
            "type": "account",
            "is_total": False,
            "values": {
                "name": name,
                "debit": 0.0,
                "credit": 0.0,
                "balance": amount,
            },
        }

    @api.model
    def _goldverse_export_filters(self, payload):
        options = payload.get("options") or {}
        filters = [
            _("Period: %s") % payload.get("period_label"),
            _("Entries: %s") % payload.get("target_label"),
            _("Currency: %s") % payload.get("currency_label"),
        ]
        if payload.get("journal_label"):
            filters.append(_("Journals: %s") % payload["journal_label"])
        if payload.get("report_key") in ("aged_receivable", "aged_payable", "aged_partner_balance"):
            filters.append(_("Aging Bucket: %s Days") % options.get("period_length", 30))
        return filters

    @api.model
    def _goldverse_build_xlsx(self, sheet, payload):
        columns = payload["export_columns"]
        column_count = max(len(columns), 1)
        last_column = get_column_letter(column_count)
        plum = "714B67"
        teal = "25B7C9"
        gold = "A36A13"
        ink = "10243A"
        muted = "5C6F84"
        line = Side(style="thin", color="D9E2EC")
        border = Border(left=line, right=line, top=line, bottom=line)
        title_fill = PatternFill("solid", fgColor=plum)
        header_fill = PatternFill("solid", fgColor=ink)
        total_fill = PatternFill("solid", fgColor="EEF3F7")
        alt_fill = PatternFill("solid", fgColor="FBFCFE")
        number_format = f'#,##0.00 "{payload.get("currency_label") or ""}";[Red]-#,##0.00 "{payload.get("currency_label") or ""}";-'

        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=column_count)
        sheet["A1"] = payload["title"]
        sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        sheet["A1"].fill = title_fill
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 28

        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=column_count)
        sheet["A2"] = payload["company"]
        sheet["A2"].font = Font(size=12, bold=True, color=ink)
        sheet["A2"].alignment = Alignment(horizontal="center")

        sheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=column_count)
        sheet["A3"] = f"{payload['period_label']} | Generated {payload['generated_at']}"
        sheet["A3"].font = Font(size=10, color=muted)
        sheet["A3"].alignment = Alignment(horizontal="center")

        sheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=column_count)
        sheet["A4"] = "   |   ".join(payload["filters"])
        sheet["A4"].font = Font(size=9, color=muted)
        sheet["A4"].alignment = Alignment(horizontal="center", wrap_text=True)

        header_row = 6
        for index, column in enumerate(columns, start=1):
            cell = sheet.cell(header_row, index, column["label"])
            cell.font = Font(size=10, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal="right" if column["type"] == "number" else "left", vertical="center")
        sheet.row_dimensions[header_row].height = 24

        for row_index, line_data in enumerate(payload["export_lines"], start=header_row + 1):
            for column_index, cell_data in enumerate(line_data["cells"], start=1):
                value = cell_data["value"]
                cell = sheet.cell(row_index, column_index)
                if cell_data["number"]:
                    cell.value = float(value or 0.0)
                    cell.number_format = number_format
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = value or ""
                    cell.alignment = Alignment(horizontal="left", indent=min(cell_data["indent"] // 14, 6))
                cell.font = Font(size=9, bold=line_data["is_total"], color=ink)
                cell.border = border
                cell.fill = total_fill if line_data["is_total"] else alt_fill if row_index % 2 else PatternFill(fill_type=None)

        for index, column in enumerate(columns, start=1):
            letter = get_column_letter(index)
            if column["type"] == "number":
                sheet.column_dimensions[letter].width = 17
            elif column["key"] in ("name", "move"):
                sheet.column_dimensions[letter].width = 38
            else:
                sheet.column_dimensions[letter].width = 18
        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.auto_filter.ref = f"A{header_row}:{last_column}{max(header_row, sheet.max_row)}"
        sheet.sheet_view.showGridLines = False
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.print_title_rows = f"{header_row}:{header_row}"

        for cell in sheet[1]:
            cell.fill = title_fill
        for row in (2, 3, 4):
            for cell in sheet[row]:
                cell.border = Border(bottom=Side(style="thin", color="E5E7EB"))

        sheet.cell(5, 1, "AimAze Laundry ERP")
        sheet.cell(5, 1).font = Font(size=9, bold=True, color=teal)
        if column_count > 1:
            sheet.cell(5, column_count, "GoldVerse Premium Laundry")
            sheet.cell(5, column_count).font = Font(size=9, bold=True, color=gold)
            sheet.cell(5, column_count).alignment = Alignment(horizontal="right")

    @api.model
    def _goldverse_safe_sheet_title(self, title):
        safe = "".join(char for char in title if char not in r"[]:*?/\\")
        return (safe or "Report")[:31]

    @api.model
    def _goldverse_export_filename(self, payload, extension):
        safe_title = "".join(char if char.isalnum() else "_" for char in payload["title"]).strip("_")
        return f"{safe_title or 'GoldVerse_Report'}_{fields.Date.today()}.{extension}"

    @api.model
    def _general_ledger_report(self, options):
        journals = self._selected_journal_ids(options)
        wizard = self.env["account.report.general.ledger"].create(
            {
                "date_from": options["date_from"],
                "date_to": options["date_to"],
                "target_move": options["target_move"],
                "display_account": options["display_account"],
                "company_id": self.env.company.id,
                "initial_balance": options["initial_balance"],
                "sortby": options["sortby"],
                "journal_ids": [(6, 0, journals)],
            }
        )
        data = {
            "ids": [],
            "model": "ir.ui.menu",
            "form": wizard.read(["date_from", "date_to", "journal_ids", "target_move", "company_id"])[0],
        }
        data["form"].update(wizard.read(["display_account", "initial_balance", "sortby"])[0])
        data["form"]["used_context"] = dict(wizard._build_contexts(data), lang=get_lang(self.env).code)
        account_domain = [("company_ids", "in", [self.env.company.id])]
        if options.get("account_ids"):
            account_domain.append(("id", "in", options["account_ids"]))
        accounts = self.env["account.account"].search(account_domain, order="code")
        account_lines = self.env["report.base_accounting_kit.report_general_ledger"].with_context(
            data["form"]["used_context"]
        )._get_account_move_entry(accounts, options["initial_balance"], options["sortby"], options["display_account"])
        columns = [
            {"key": "name", "label": _(""), "type": "text"},
            {"key": "date", "label": _("Date"), "type": "text"},
            {"key": "partner", "label": _("Partner"), "type": "text"},
            {"key": "currency", "label": _("Currency"), "type": "text"},
            {"key": "debit", "label": _("Debit"), "type": "number"},
            {"key": "credit", "label": _("Credit"), "type": "number"},
            {"key": "balance", "label": _("Balance"), "type": "number"},
        ]
        lines = []
        accounts_by_code = {account.code: account.id for account in accounts}
        line_ids = [
            move.get("lid")
            for account in account_lines
            for move in (account.get("move_lines") or [])
            if move.get("lid")
        ]
        move_by_line_id = {
            line.id: line.move_id.id
            for line in self.env["account.move.line"].browse(line_ids).exists()
            if line.move_id
        }
        for index, account in enumerate(account_lines):
            real_account_id = accounts_by_code.get(account["code"])
            account_id = f"account_{real_account_id or index}"
            lines.append(
                {
                    "id": account_id,
                    "name": f"{account['code']} {account['name']}",
                    "level": 1,
                    "type": "account",
                    "is_total": True,
                    "action": self._move_line_action(
                        _("Journal Items"),
                        options,
                        account_ids=[real_account_id],
                        date_from=False,
                    )
                    if real_account_id
                    else False,
                    "values": {
                        "date": "",
                        "name": f"{account['code']} {account['name']}",
                        "journal": "",
                        "partner": "",
                        "currency": self.env.company.currency_id.name or "",
                        "debit": account.get("debit") or 0.0,
                        "credit": account.get("credit") or 0.0,
                        "balance": account.get("balance") or 0.0,
                    },
                }
            )
            for move_index, move in enumerate(account.get("move_lines") or []):
                move_id = move_by_line_id.get(move.get("lid"))
                move_name = move.get("move_name") or ""
                line_name = move.get("lname") or ""
                if move_name and line_name and line_name not in move_name:
                    entry_name = f"{move_name} {line_name}"
                else:
                    entry_name = line_name or move_name
                lines.append(
                    {
                        "id": f"{account_id}_{move_index}",
                        "name": entry_name,
                        "level": 2,
                        "type": "move",
                        "is_total": False,
                        "move_action": self._journal_entry_action(move_id) if move_id else False,
                        "move_url": f"/odoo/action-274/{move_id}" if move_id else False,
                        "values": {
                            "date": fields.Date.to_string(move.get("ldate")) if move.get("ldate") else "",
                            "name": entry_name,
                            "journal": move.get("lcode") or "",
                            "partner": move.get("partner_name") or "",
                            "currency": self.env.company.currency_id.name or "",
                            "debit": move.get("debit") or 0.0,
                            "credit": move.get("credit") or 0.0,
                            "balance": move.get("balance") or 0.0,
                        },
                    }
                )
        return columns, lines

    @api.model
    def _aged_partner_report(self, options):
        account_types, report_name = self._goldverse_aged_account_types(options)
        date_to = fields.Date.from_string(options["date_to"])
        period_length = max(options["period_length"], 1)
        currency = self.env.company.currency_id
        periods = {
            "4": f"1-{period_length}",
            "3": f"{period_length + 1}-{period_length * 2}",
            "2": f"{period_length * 2 + 1}-{period_length * 3}",
            "1": f"{period_length * 3 + 1}-{period_length * 4}",
            "0": _("Older"),
        }
        columns = [
            {"key": "name", "label": "", "type": "text"},
            {"key": "invoice_date", "label": _("Invoice Date"), "type": "text"},
            {"key": "direction", "label": _("At Date"), "type": "number"},
            {"key": "4", "label": periods["4"], "type": "number"},
            {"key": "3", "label": periods["3"], "type": "number"},
            {"key": "2", "label": periods["2"], "type": "number"},
            {"key": "1", "label": periods["1"], "type": "number"},
            {"key": "0", "label": periods["0"], "type": "number"},
            {"key": "total", "label": _("Total"), "type": "number"},
        ]
        bucket_keys = ("direction", "4", "3", "2", "1", "0", "total")
        totals = {key: 0.0 for key in bucket_keys}
        partners = {}
        move_states = ["posted"] if options["target_move"] == "posted" else ["draft", "posted"]
        domain = [
            ("company_id", "=", self.env.company.id),
            ("parent_state", "in", move_states),
            ("account_id.account_type", "in", account_types),
            ("date", "<=", date_to),
        ]
        move_lines = self.env["account.move.line"].sudo().search(
            domain,
            order="partner_id, date_maturity, date, id",
        )
        for line in move_lines:
            amount = self._goldverse_aged_open_amount(line, date_to, currency)
            if currency.is_zero(amount):
                continue
            partner_id = line.partner_id.id or False
            partner_values = partners.setdefault(
                partner_id,
                {
                    "partner_id": partner_id,
                    "name": line.partner_id.display_name or _("Unknown Partner"),
                    **{key: 0.0 for key in bucket_keys},
                },
            )
            bucket = self._goldverse_aged_bucket(line.date_maturity or line.date, date_to, period_length)
            partner_values[bucket] += amount
            partner_values["total"] += amount
            totals[bucket] += amount
            totals["total"] += amount
        lines = []
        sorted_partners = sorted(
            (
                partner_line
                for partner_line in partners.values()
                if not currency.is_zero(partner_line["total"])
            ),
            key=lambda value: (value["name"] or "").casefold(),
        )
        for index, partner_line in enumerate(sorted_partners):
            values = {
                "name": partner_line["name"],
                "invoice_date": "",
            }
            values.update({key: partner_line[key] for key in bucket_keys})
            action = False
            if partner_line["partner_id"]:
                action = self._goldverse_aged_ledger_action(
                    partner_line["name"] or _("Partner Ledger"),
                    options,
                    account_types,
                    partner_id=partner_line["partner_id"],
                )
            lines.append(
                {
                    "id": f"aged_{index}",
                    "name": values["name"],
                    "level": 2,
                    "type": "partner",
                    "is_total": False,
                    "action": action,
                    "values": values,
                }
            )
        total_label = _("Total Receivable") if options["result_selection"] == "customer" else report_name
        if options["result_selection"] == "supplier":
            total_label = _("Total Payable")
        lines.append(
            {
                "id": "aged_total",
                "name": total_label,
                "level": 1,
                "type": "aged_total",
                "is_total": True,
                "action": self._goldverse_aged_ledger_action(total_label, options, account_types),
                "values": {
                    "name": total_label,
                    "invoice_date": "",
                    **totals,
                },
            }
        )
        return columns, lines

    @api.model
    def _goldverse_aged_ledger_action(self, name, options, account_types, partner_id=False):
        domain = self._move_line_domain(
            options,
            account_types=account_types,
            date_from=False,
        )
        if partner_id:
            domain.append(("partner_id", "=", partner_id))
        list_view = self.env.ref(
            "goldverse_premium_laundry_branding.view_goldverse_aged_partner_ledger_line_list",
            raise_if_not_found=False,
        )
        return {
            "type": "ir.actions.act_window",
            "name": name,
            "res_model": "account.move.line",
            "view_mode": "list,form",
            "views": [(list_view.id, "list"), (False, "form")] if list_view else [(False, "list"), (False, "form")],
            "domain": domain,
            "context": {
                "default_company_id": self.env.company.id,
            },
        }

    @api.model
    def _goldverse_aged_account_types(self, options):
        if options["result_selection"] == "customer":
            return ["asset_receivable"], _("Aged Receivable")
        if options["result_selection"] == "supplier":
            return ["liability_payable"], _("Aged Payable")
        return ["liability_payable", "asset_receivable"], _("Aged Partner Balance")

    @api.model
    def _goldverse_aged_bucket(self, maturity_date, date_to, period_length):
        if maturity_date >= date_to:
            return "direction"
        days_overdue = (date_to - maturity_date).days
        if days_overdue <= period_length:
            return "4"
        if days_overdue <= period_length * 2:
            return "3"
        if days_overdue <= period_length * 3:
            return "2"
        if days_overdue <= period_length * 4:
            return "1"
        return "0"

    @api.model
    def _goldverse_aged_open_amount(self, line, date_to, currency):
        """Return the line's residual amount as of date_to in `currency`.

        account.move.line.matched_debit_ids holds partial reconciles where
        THIS line is on the credit side (so the matched partner is a debit
        that offsets it) — to compute the residual we ADD those debits to
        the balance (a credit balance like -1000 + matched debit 1000 = 0).
        Conversely matched_credit_ids holds partials where THIS line is on
        the debit side, so we SUBTRACT the matched credits (invoice +1000
        - matched credit 1000 = 0).

        The previous implementation had these signs swapped, which doubled
        the absolute balance of every reconciled line in each aging bucket
        (per-partner totals still cancelled to the right residual but the
        per-bucket "Total Receivable" row was inflated into the millions).
        """
        company_currency = line.company_id.currency_id
        amount = company_currency._convert(line.balance, currency, line.company_id, date_to)
        for partial in line.matched_debit_ids:
            if partial.max_date <= date_to:
                amount += partial.company_id.currency_id._convert(
                    partial.amount,
                    currency,
                    partial.company_id,
                    date_to,
                )
        for partial in line.matched_credit_ids:
            if partial.max_date <= date_to:
                amount -= partial.company_id.currency_id._convert(
                    partial.amount,
                    currency,
                    partial.company_id,
                    date_to,
                )
        return amount

    @api.model
    def _journal_entry_action(self, move_id):
        return {
            "type": "ir.actions.act_window",
            "name": _("Journal Voucher"),
            "res_model": "account.move",
            "res_id": move_id,
            "view_mode": "form",
            "views": [(False, "form")],
            "target": "current",
        }


class GoldverseInteractiveAccountReportPdf(models.AbstractModel):
    _name = "report.goldverse_premium_laundry_branding.gv_account_pdf"
    _description = "GoldVerse Interactive Accounting Report PDF"

    @api.model
    def _get_report_values(self, docids, data=None):
        data = data or {}
        payload = self.env["account.interactive.report"].goldverse_export_payload(
            data.get("report_key") or "profit_and_loss",
            data.get("options") or {},
        )
        return {
            "doc_ids": docids,
            "doc_model": "res.company",
            "docs": self.env["res.company"].browse(docids),
            "payload": payload,
        }
