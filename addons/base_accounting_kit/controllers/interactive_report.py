# -*- coding: utf-8 -*-
import json
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from odoo import http
from odoo.http import content_disposition, request


class InteractiveAccountReportController(http.Controller):
    @http.route(
        "/base_accounting_kit/interactive_report/xlsx",
        type="http",
        auth="user",
        methods=["POST"],
        csrf=False,
    )
    def interactive_report_xlsx(self, report_key=None, options=None, **_kw):
        report_key = report_key or "profit_and_loss"
        parsed_options = json.loads(options or "{}")
        payload = request.env["account.interactive.report"].get_report(report_key, parsed_options)

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = payload["title"][:31]
        if report_key == "trial_balance":
            self._build_trial_balance_xlsx(sheet, payload)
            output = BytesIO()
            workbook.save(output)
            filename = f"{payload['title'].replace(' ', '_')}.xlsx"
            headers = [
                ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
                ("Content-Disposition", content_disposition(filename)),
            ]
            return request.make_response(output.getvalue(), headers=headers)

        sheet.append([payload["title"]])
        sheet.append([payload["company"]])
        sheet.append([])
        sheet.append([column["label"] for column in payload["columns"]])

        for line in payload["lines"]:
            row = []
            for column in payload["columns"]:
                value = line["values"].get(column["key"], "")
                if column["key"] in ("name", "move") and line.get("level", 1) > 1:
                    value = f"{'  ' * (line['level'] - 1)}{value}"
                row.append(value)
            sheet.append(row)

        for column_cells in sheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 45)

        output = BytesIO()
        workbook.save(output)
        filename = f"{payload['title'].replace(' ', '_')}.xlsx"
        headers = [
            ("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            ("Content-Disposition", content_disposition(filename)),
        ]
        return request.make_response(output.getvalue(), headers=headers)

    def _build_trial_balance_xlsx(self, sheet, payload):
        thin = Side(style="thin", color="D5DBE3")
        border = Border(bottom=thin)
        box_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        title_fill = PatternFill("solid", fgColor="FFFFFF")
        current_fill = PatternFill("solid", fgColor="F8FAFC")
        muted_fill = PatternFill("solid", fgColor="F3F4F6")
        sheet.append([payload["title"]])
        sheet["A1"].font = Font(size=16, bold=True, color="111827")
        sheet.append([payload["company"], "", "", "", f"In {payload['currency_label']}"])
        sheet.append([])
        sheet.append(["", "Initial Balance", self._period_label(payload["options"]), "", "End Balance"])
        sheet.merge_cells(start_row=4, start_column=3, end_row=4, end_column=4)
        sheet.append(["", "Balance", "Debit", "Credit", "Balance"])

        for cell in sheet[4]:
            cell.font = Font(bold=True, color="111827")
            cell.alignment = Alignment(horizontal="center")
            if cell.column > 1:
                cell.border = box_border
                cell.fill = title_fill
        for cell in sheet[5]:
            cell.font = Font(bold=True, color="111827")
            cell.alignment = Alignment(horizontal="right" if cell.column > 1 else "left")
            cell.border = border
            if cell.column in (3, 4):
                cell.fill = current_fill

        for line in payload["lines"]:
            values = line.get("values", {})
            row = [
                values.get("name", ""),
                values.get("initial_balance") or 0.0,
                values.get("debit") or 0.0,
                values.get("credit") or 0.0,
                values.get("end_balance") or 0.0,
            ]
            sheet.append(row)
            excel_row = sheet.max_row
            for index, cell in enumerate(sheet[excel_row], start=1):
                cell.border = border
                cell.alignment = Alignment(horizontal="right" if index > 1 else "left")
                if index in (3, 4):
                    cell.fill = current_fill
                if index > 1:
                    cell.number_format = '#,##0.00;[Red]-#,##0.00;#,##0.00'
                if line.get("is_total"):
                    cell.font = Font(bold=True, color="111827")
                    cell.fill = muted_fill if index > 1 else title_fill

        sheet.column_dimensions["A"].width = 44
        for column in ("B", "C", "D", "E"):
            sheet.column_dimensions[column].width = 17
        sheet.freeze_panes = "A6"

    def _period_label(self, options):
        period = options.get("period")
        date_from = options.get("date_from")
        date_to = options.get("date_to")
        try:
            start = datetime.strptime(date_from, "%Y-%m-%d")
            end = datetime.strptime(date_to, "%Y-%m-%d")
        except Exception:
            return "%s to %s" % (date_from or "", date_to or "")
        if period == "month":
            return start.strftime("%B %Y")
        if period == "quarter":
            return "%s - %s %s" % (start.strftime("%b"), end.strftime("%b"), end.year)
        if period == "year":
            return str(start.year)
        return "%s to %s" % (date_from, date_to)
