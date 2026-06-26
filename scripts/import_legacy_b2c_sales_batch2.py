"""Import GoldVerse legacy B2C Batch-2 (Unpaid) workbook.

Run through Odoo shell. Path read from GOLDVERSE_LEGACY_B2C_BATCH2_XLSX.
By default this performs a rollback dry-run. Set
GOLDVERSE_LEGACY_B2C_BATCH2_COMMIT=1 to commit.

Differences vs Batch-1:
- Real customer data: match res.partner by last-10-digits of mobile,
  create new B2C partner if missing.
- Most rows are unpaid: invoice posted, no payment, AR open.
- The one paid row (GPL/EME/POS/112301) gets a Cash payment registered
  for its workbook Cash Received amount and is reconciled to zero AR.
- Orders are marked delivered (per user direction).
- Idempotency marker: GOLDVERSE_LEGACY_B2C_SALES_IMPORT_BATCH2.
"""

from collections import OrderedDict
from datetime import datetime, time
import json
import os
from pathlib import Path

import pytz
from openpyxl import load_workbook

from odoo import api, fields
from odoo.exceptions import UserError, ValidationError


IMPORT_MARKER = "GOLDVERSE_LEGACY_B2C_SALES_IMPORT_BATCH2"
COMPANY_NAME = "GoldVerse Premium (Pvt.) Limited"
LEGACY_CATEGORY_NAME = "Legacy Import"
LEGACY_SERVICE_NAME = "Legacy Imported B2C Sale"
DEFAULT_SOURCE = "walk_in"
DELIVERY_TZ = "Asia/Karachi"


def _money(value):
    return round(float(value or 0.0), 2)


def _float_or_zero(value):
    if value in (None, False, ""):
        return 0.0
    return float(value)


def _clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _clean_mobile(value):
    if value is None:
        return ""
    digits = "".join(c for c in str(value) if c.isdigit())
    return digits[-10:] if len(digits) >= 10 else digits


def _date_or_false(value):
    if not value:
        return False
    if isinstance(value, datetime):
        return value.date()
    return fields.Date.to_date(value)


def _datetime_or_false(value):
    if not value:
        return False
    if isinstance(value, datetime):
        return fields.Datetime.to_datetime(value)
    return fields.Datetime.to_datetime(value)


def _legacy_delivery_datetime(value):
    base_dt = _datetime_or_false(value)
    if not base_dt:
        return False
    delivery_tz = pytz.timezone(DELIVERY_TZ)
    if base_dt.tzinfo:
        localized = base_dt.astimezone(delivery_tz)
    else:
        localized = delivery_tz.localize(base_dt)
    target_local = delivery_tz.localize(datetime.combine(localized.date(), time(18, 0)))
    utc_dt = target_local.astimezone(pytz.UTC).replace(tzinfo=None)
    return fields.Datetime.to_string(utc_dt)


def _load_rows(xlsx_path):
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    grouped_orders = OrderedDict()
    totals = {
        "amount_total": 0.0,
        "discount": 0.0,
        "net_total": 0.0,
        "cash_received": 0.0,
        "ibft_received": 0.0,
        "balance_due": 0.0,
    }
    skipped_rows = []

    for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        order_name = _clean_text(values[5])
        if not order_name:
            continue
        payment_status = (_clean_text(values[15]) or "Unpaid").lower()
        row = {
            "row_index": row_index,
            "branch": _clean_text(values[0]) or "EME",
            "customer_type": (_clean_text(values[1]) or "B2C").upper(),
            "mobile_raw": values[2],
            "mobile_clean": _clean_mobile(values[2]),
            "customer_name": _clean_text(values[3]),
            "order_date": values[4],
            "order_name": order_name,
            "invoice_name": _clean_text(values[6]),
            "total_qty": _float_or_zero(values[7]),
            "amount_total": _money(values[8]),
            "discount": _money(values[9]),
            "net_total": _money(values[10]),
            "payment_date": values[11] or values[4],
            "cash_received": _money(values[12]),
            "ibft_received": _money(values[13]),
            "balance_due": _money(values[14]),
            "payment_status": _clean_text(values[15]) or "Unpaid",
            "state": _clean_text(values[16]) or "Pending Delivery to Customer",
            "source_file": str(xlsx_path),
        }
        for key in totals:
            totals[key] += row[key]
        # One order per row (no grouping in this workbook), but wrap as list for parity with Batch-1
        grouped_orders[order_name] = {
            "order_name": order_name,
            "branch": row["branch"],
            "customer_type": row["customer_type"],
            "mobile_raw": row["mobile_raw"],
            "mobile_clean": row["mobile_clean"],
            "customer_name": row["customer_name"],
            "order_date": row["order_date"],
            "state": row["state"],
            "source_file": row["source_file"],
            "row_indexes": [row_index],
            "lines": [row],
            "amount_total": row["amount_total"],
            "discount": row["discount"],
            "net_total": row["net_total"],
            "cash_received": row["cash_received"],
            "ibft_received": row["ibft_received"],
            "balance_due": row["balance_due"],
        }

    return {
        "rows": list(grouped_orders.values()),
        "totals": {key: _money(value) for key, value in totals.items()},
        "skipped_rows": skipped_rows,
    }


def _ensure_record(model, domain, values):
    record = model.search(domain, limit=1)
    if record:
        return record
    return model.create(values)


def _prepare_env(env):
    company = env["res.company"].sudo().search([("name", "=", COMPANY_NAME)], limit=1)
    if not company:
        raise UserError("Missing company for legacy import: %s" % COMPANY_NAME)

    ctx = dict(env.context, allowed_company_ids=[company.id], force_company=company.id)
    company_env = api.Environment(env.cr, env.uid, ctx)

    Partner = company_env["res.partner"].sudo()
    Category = company_env["aimaze.laundry.service.category"].sudo()
    Service = company_env["aimaze.laundry.service"].sudo()
    Colour = company_env["goldverse.laundry.colour"].sudo()
    Topup = company_env["goldverse.laundry.topup"].sudo()
    Branch = company_env["aimaze.laundry.branch"].sudo()
    Journal = company_env["account.journal"].sudo()

    category = _ensure_record(Category, [("name", "=", LEGACY_CATEGORY_NAME)], {"name": LEGACY_CATEGORY_NAME})
    colour = (
        Colour.search([("name", "ilike", "multi")], limit=1)
        or Colour.search([], limit=1)
        or _ensure_record(Colour, [("name", "=", "Legacy Unknown")], {"name": "Legacy Unknown"})
    )
    topup = (
        Topup.search([("name", "=", "Folded Pack")], limit=1)
        or Topup.search([], limit=1)
    )
    if not topup:
        raise UserError("Missing Add On records for legacy import.")

    service = Service.search([("name", "=", LEGACY_SERVICE_NAME)], limit=1)
    if not service:
        service = Service.create(
            {
                "name": LEGACY_SERVICE_NAME,
                "category_id": category.id,
                "list_price": 0.0,
                "goldverse_base_price": 0.0,
            }
        )

    branch = (
        Branch.search([("name", "=", "EME Branch")], limit=1)
        or Branch.search([("name", "ilike", "EME")], limit=1)
        or Branch.search([], limit=1)
    )
    if not branch:
        raise UserError("Missing EME branch for legacy import.")

    cash_journal = Journal.search([("company_id", "=", company.id), ("name", "=", "Cash")], limit=1)
    ibft_journal = Journal.search([("company_id", "=", company.id), ("name", "=", "IBFT")], limit=1)
    if not cash_journal or not cash_journal.inbound_payment_method_line_ids:
        raise UserError("Missing Cash journal or inbound payment method line for legacy import.")
    if not ibft_journal or not ibft_journal.inbound_payment_method_line_ids:
        raise UserError("Missing IBFT journal or inbound payment method line for legacy import.")

    return {
        "env": company_env,
        "company": company,
        "category": category,
        "service": service,
        "colour": colour,
        "topup": topup,
        "branch": branch,
        "Partner": Partner,
        "cash_journal": cash_journal,
        "cash_method_line": cash_journal.inbound_payment_method_line_ids[:1],
        "ibft_journal": ibft_journal,
        "ibft_method_line": ibft_journal.inbound_payment_method_line_ids[:1],
    }


def _resolve_partner(prepared, row, partner_cache, partner_stats):
    Partner = prepared["Partner"]
    mob_clean = row["mobile_clean"]
    name = row["customer_name"] or "Unknown Customer"
    if mob_clean and mob_clean in partner_cache:
        return partner_cache[mob_clean]
    partner = False
    if mob_clean:
        # Search candidates whose stripped mobile ends with our cleaned digits
        candidates = Partner.search([("mobile", "!=", False)], limit=5000)
        for cand in candidates:
            if _clean_mobile(cand.mobile) == mob_clean:
                partner = cand
                break
        if not partner:
            candidates = Partner.search([("phone", "!=", False)], limit=5000)
            for cand in candidates:
                if _clean_mobile(cand.phone) == mob_clean:
                    partner = cand
                    break
    if partner:
        partner_stats["matched_existing"] += 1
    else:
        full_mobile = "0" + mob_clean if mob_clean and len(mob_clean) == 10 else (mob_clean or "")
        partner = Partner.with_context(goldverse_skip_duplicate_mobile_check=True).create(
            {
                "name": name,
                "mobile": full_mobile or False,
                "phone": full_mobile or False,
                "customer_rank": 1,
                "company_id": prepared["company"].id,
                "goldverse_customer_category": "b2c",
                "laundry_customer_type": "b2c",
            }
        )
        partner_stats["created_new"] += 1
    if mob_clean:
        partner_cache[mob_clean] = partner
    return partner


def _validate_rows(rows):
    problems = []
    for row in rows:
        if abs(row["amount_total"] - row["discount"] - row["net_total"]) > 0.01:
            problems.append(
                "Order %(order_name)s amount %(amount_total).2f - discount %(discount).2f != net %(net_total).2f" % row
            )
        if not row["customer_name"]:
            problems.append("Order %(order_name)s has empty customer name" % row)
        if not row["mobile_clean"]:
            problems.append("Order %(order_name)s has empty/invalid mobile" % row)
        if row["customer_type"] != "B2C":
            problems.append("Order %(order_name)s customer type is %(customer_type)s" % row)
    if problems:
        raise ValidationError("\n".join(problems[:50]))


def _existing_order_names(order_model):
    return set(order_model.search([]).mapped("name"))


def _reconcile_payment_to_invoice(invoice, payment):
    invoice_lines = invoice.line_ids.filtered(
        lambda line: line.account_id.account_type == "asset_receivable" and not line.reconciled
    )
    payment_lines = payment.move_id.line_ids.filtered(
        lambda line: line.account_id.account_type == "asset_receivable" and not line.reconciled
    )
    (invoice_lines + payment_lines).reconcile()


def _create_payment(prepared, partner, order, invoice, amount, payment_date, journal, method_line, memo):
    if amount <= 0:
        return False
    receivable_line = invoice.line_ids.filtered(
        lambda line: line.account_id.account_type == "asset_receivable"
    )[:1]
    payment = prepared["env"]["account.payment"].sudo().create(
        {
            "company_id": prepared["company"].id,
            "payment_type": "inbound",
            "partner_type": "customer",
            "partner_id": partner.id,
            "aimaze_laundry_order_id": order.id,
            "amount": amount,
            "date": payment_date,
            "journal_id": journal.id,
            "payment_method_line_id": method_line.id,
            "memo": memo,
            "destination_account_id": receivable_line.account_id.id if receivable_line else False,
        }
    )
    payment.action_post()
    _reconcile_payment_to_invoice(invoice, payment)
    return payment


def _patch_order_report_fields(order, row):
    values = {
        "goldverse_report_base_price": row["amount_total"],
        "goldverse_report_discount": row["discount"],
        "goldverse_report_net_price": row["net_total"],
        "goldverse_report_priority_charges": 0.0,
        "goldverse_report_total_excl_tax": row["net_total"],
        "goldverse_report_payment_date": _date_or_false(row["lines"][0]["payment_date"]) if row.get("cash_received") or row.get("ibft_received") else False,
        "goldverse_report_cash_received": row.get("cash_received") or 0.0,
        "goldverse_report_ibft_received": row.get("ibft_received") or 0.0,
        "delivery_charge": 0.0,
    }
    order.with_context(
        goldverse_allow_locked_order_write=True,
        goldverse_skip_required_validation=True,
    ).write(values)


def _patch_order_line_breakdown(order, row):
    ordered_lines = order.line_ids.sorted("id")
    for line, line_row in zip(ordered_lines, row["lines"]):
        line.with_context(
            goldverse_allow_locked_order_write=True,
            goldverse_refreshing_amounts=True,
        ).write(
            {
                "goldverse_legacy_amount_override": True,
                "goldverse_legacy_base_unit_price": line_row["amount_total"],
                "goldverse_legacy_discount_amount": line_row["discount"],
                "goldverse_legacy_net_unit_price": line_row["net_total"],
                "goldverse_legacy_priority_charge": 0.0,
                "unit_price": line_row["net_total"],
                "discount": 0.0,
                "goldverse_discount": "0",
                "tax_ids": [(5, 0, 0)],
                "goldverse_base_price": line_row["amount_total"],
                "goldverse_discount_amount": line_row["discount"],
                "goldverse_net_price": line_row["net_total"],
                "goldverse_priority_charge": 0.0,
            }
        )


def _import_rows(env, workbook_path, commit=False, limit=0):
    payload = _load_rows(workbook_path)
    rows = payload["rows"]
    _validate_rows(rows)
    prepared = _prepare_env(env)
    Order = prepared["env"]["aimaze.laundry.order"].sudo()
    existing_order_names = _existing_order_names(Order)
    chunk_size = int(os.environ.get("GOLDVERSE_LEGACY_B2C_BATCH2_CHUNK_SIZE") or 200)
    selected_rows = rows[:limit or None]

    created_orders = 0
    created_invoices = 0
    created_payments = 0
    skipped = 0
    imported_amounts = {
        "amount_total": 0.0,
        "discount": 0.0,
        "net_total": 0.0,
        "cash_received": 0.0,
        "ibft_received": 0.0,
    }
    partner_cache = {}
    partner_stats = {"matched_existing": 0, "created_new": 0}
    first_invoice_name = False
    last_invoice_name = False
    per_order_log = []
    paid_orders_log = []

    for chunk_start in range(0, len(selected_rows), chunk_size):
        chunk_rows = selected_rows[chunk_start : chunk_start + chunk_size]
        for row in chunk_rows:
            if row["order_name"] in existing_order_names:
                skipped += 1
                per_order_log.append({"order": row["order_name"], "action": "skipped_existing"})
                continue

            partner = _resolve_partner(prepared, row, partner_cache, partner_stats)

            order_date_dt = _datetime_or_false(row["order_date"]) or fields.Datetime.now()
            order_date_str = fields.Datetime.to_string(order_date_dt)
            delivery_datetime_str = _legacy_delivery_datetime(row["order_date"]) or order_date_str
            line_commands = [
                (
                    0,
                    0,
                    {
                        "service_id": prepared["service"].id,
                        "quantity": 1,
                        "goldverse_priority": "normal",
                        "goldverse_category_id": prepared["category"].id,
                        "goldverse_colour_ids": [(6, 0, [prepared["colour"].id])],
                        "goldverse_topup_ids": [(6, 0, [prepared["topup"].id])],
                        "name": "%s #1" % LEGACY_SERVICE_NAME,
                    },
                )
            ]

            order = Order.create(
                {
                    "name": row["order_name"],
                    "barcode": row["order_name"],
                    "partner_id": partner.id,
                    "mobile_partner_id": partner.id,
                    "mobile": partner.mobile,
                    "email": False,
                    "customer_type": "b2c",
                    "source": DEFAULT_SOURCE,
                    "priority": "normal",
                    "branch_id": prepared["branch"].id,
                    "order_date": order_date_str,
                    "expected_delivery_datetime": delivery_datetime_str,
                    "line_ids": line_commands,
                }
            )

            for line in order.line_ids.sorted("id"):
                line.with_context(goldverse_refreshing_amounts=True).write(
                    {
                        "quantity": 1,
                        "unit_price": row["net_total"],
                        "discount": 0.0,
                        "tax_ids": [(5, 0, 0)],
                    }
                )

            order.action_create_order()
            order.action_create_invoice()
            invoice = order.invoice_id
            invoice.write(
                {
                    "invoice_date": _date_or_false(row["order_date"]),
                    "ref": row["order_name"],
                    "invoice_origin": row["order_name"],
                    "payment_reference": row["order_name"],
                    "narration": "%s\nSource file row: %s\nWorkbook: %s"
                    % (IMPORT_MARKER, row["row_indexes"][0], row["source_file"]),
                }
            )
            if invoice.state == "draft":
                invoice.action_post()
            created_invoices += 1
            if not first_invoice_name:
                first_invoice_name = invoice.name
            last_invoice_name = invoice.name

            # Register payments if the workbook row reports cash/IBFT received.
            payment_date = _date_or_false(row["lines"][0]["payment_date"]) or _date_or_false(row["order_date"])
            if row.get("cash_received") and row["cash_received"] > 0:
                cash_payment = _create_payment(
                    prepared,
                    partner,
                    order,
                    invoice,
                    row["cash_received"],
                    payment_date,
                    prepared["cash_journal"],
                    prepared["cash_method_line"],
                    "%s CASH" % row["order_name"],
                )
                if cash_payment:
                    created_payments += 1
                    paid_orders_log.append({
                        "order": row["order_name"],
                        "partner": partner.name,
                        "cash": row["cash_received"],
                        "payment_name": cash_payment.name,
                    })
            if row.get("ibft_received") and row["ibft_received"] > 0:
                ibft_payment = _create_payment(
                    prepared,
                    partner,
                    order,
                    invoice,
                    row["ibft_received"],
                    payment_date,
                    prepared["ibft_journal"],
                    prepared["ibft_method_line"],
                    "%s IBFT" % row["order_name"],
                )
                if ibft_payment:
                    created_payments += 1

            # Mark delivered (per user direction). Order state goes to 'paid'
            # when an inbound payment was registered, else 'delivered'.
            order.with_context(
                goldverse_allow_locked_order_write=True,
                goldverse_skip_required_validation=True,
            ).write(
                {
                    "warehouse_collected_datetime": order_date_str,
                    "warehouse_received_datetime": order_date_str,
                    "goldverse_delivered_to_customer": True,
                    "goldverse_actual_delivery_datetime": order_date_str,
                }
            )
            target_state = "paid" if (row.get("cash_received") or row.get("ibft_received")) else "delivered"
            order.with_context(
                goldverse_allow_locked_order_write=True,
                goldverse_skip_required_validation=True,
            )._set_state(target_state)
            _patch_order_line_breakdown(order, row)
            _patch_order_report_fields(order, row)

            created_orders += 1
            existing_order_names.add(row["order_name"])
            for key in imported_amounts:
                imported_amounts[key] += row[key]
            per_order_log.append({
                "order": row["order_name"],
                "partner": partner.name,
                "mobile": partner.mobile,
                "net": row["net_total"],
            })

        if commit:
            env.cr.commit()
        print(
            json.dumps(
                {
                    "chunk_start": chunk_start + 1,
                    "chunk_end": min(chunk_start + len(chunk_rows), len(selected_rows)),
                    "created_orders": created_orders,
                    "created_invoices": created_invoices,
                    "skipped_existing_orders": skipped,
                    "partner_stats": partner_stats,
                },
                sort_keys=True,
            ),
            flush=True,
        )

    result = {
        "commit": bool(commit),
        "validation_only": not commit,
        "created_orders": created_orders,
        "created_invoices": created_invoices,
        "created_payments": created_payments,
        "skipped_existing_orders": skipped,
        "workbook_rows_processed": len(selected_rows),
        "workbook_totals": payload["totals"],
        "imported_totals": {key: _money(value) for key, value in imported_amounts.items()},
        "first_invoice_name": first_invoice_name,
        "last_invoice_name": last_invoice_name,
        "partner_stats": partner_stats,
        "paid_orders": paid_orders_log,
        "sample_orders": per_order_log[:10],
    }
    if commit:
        env.cr.commit()
        result["transaction"] = "committed"
    else:
        env.cr.rollback()
        result["transaction"] = "rolled_back"
    return result


workbook_env = os.environ.get("GOLDVERSE_LEGACY_B2C_BATCH2_XLSX")
if not workbook_env:
    raise UserError("Set GOLDVERSE_LEGACY_B2C_BATCH2_XLSX before running this script.")

workbook_path = Path(workbook_env)
if not workbook_path.exists():
    raise UserError("Workbook not found: %s" % workbook_path)

limit = int(os.environ.get("GOLDVERSE_LEGACY_B2C_BATCH2_LIMIT") or 0)
commit = os.environ.get("GOLDVERSE_LEGACY_B2C_BATCH2_COMMIT") == "1"
result = _import_rows(env, workbook_path, commit=commit, limit=limit)
print(json.dumps(result, indent=2, sort_keys=True, default=str))
