"""Import GoldVerse legacy B2C sales workbook into the GoldVerse tenant.

Run through Odoo shell. The workbook path is read from
GOLDVERSE_LEGACY_B2C_XLSX. By default this performs a rollback dry-run.
Set GOLDVERSE_LEGACY_B2C_COMMIT=1 to commit.

What this import does:
- preserves the old Order No. from the workbook as the laundry order name
- creates a clearly-labeled legacy B2C customer/service bucket because the
  workbook does not contain real customer names, mobiles, or service items
- groups duplicate Order No. rows into one imported legacy order
- creates and posts customer invoices
- registers and posts Cash / IBFT payments using the workbook payment dates
- marks the imported orders as delivered/paid
- patches stored report helper fields so GoldVerse reports match workbook totals
"""

from collections import OrderedDict
from datetime import datetime, time
import json
import os
from pathlib import Path

import pytz
from openpyxl import load_workbook

from odoo import api, fields
from odoo.exceptions import UserError, ValidationError


IMPORT_MARKER = "GOLDVERSE_LEGACY_B2C_SALES_IMPORT"
COMPANY_NAME = "GoldVerse Premium (Pvt.) Limited"
LEGACY_CUSTOMER_NAME = "Legacy Imported B2C Sales"
LEGACY_CUSTOMER_MOBILE = "00000000000"
LEGACY_CATEGORY_NAME = "Legacy Import"
LEGACY_SERVICE_NAME = "Legacy Imported B2C Sale"
DEFAULT_SOURCE = "walk_in"
DELIVERY_TZ = "Asia/Karachi"


def _money(value):
    return round(float(value or 0.0), 2)


def _float_or_zero(value):
    if value in (None, False, ""):
        return 0.0
    return float(value)


def _clean_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _date_or_false(value):
    if not value:
        return False
    if isinstance(value, datetime):
        return value.date()
    return fields.Date.to_date(value)


def _datetime_or_false(value):
    if not value:
        return False
    if isinstance(value, datetime):
        return fields.Datetime.to_datetime(value)
    return fields.Datetime.to_datetime(value)


def _legacy_delivery_datetime(value):
    base_dt = _datetime_or_false(value)
    if not base_dt:
        return False
    delivery_tz = pytz.timezone(DELIVERY_TZ)
    if base_dt.tzinfo:
        localized = base_dt.astimezone(delivery_tz)
    else:
        localized = delivery_tz.localize(base_dt)
    target_local = delivery_tz.localize(datetime.combine(localized.date(), time(18, 0)))
    utc_dt = target_local.astimezone(pytz.UTC).replace(tzinfo=None)
    return fields.Datetime.to_string(utc_dt)


def _payment_datetime(value):
    base_dt = _datetime_or_false(value)
    if not base_dt:
        return fields.Datetime.now()
    if base_dt.tzinfo:
        return fields.Datetime.to_string(base_dt.astimezone(pytz.UTC).replace(tzinfo=None))
    return fields.Datetime.to_string(base_dt)


def _load_rows(xlsx_path):
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    grouped_orders = OrderedDict()
    totals = {
        "amount_total": 0.0,
        "discount": 0.0,
        "net_total": 0.0,
        "cash_received": 0.0,
        "ibft_received": 0.0,
        "balance_due": 0.0,
    }

    for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        order_name = _clean_text(values[5])
        if not order_name:
            continue
        row = {
            "row_index": row_index,
            "branch": _clean_text(values[0]) or "EME",
            "customer_type": (_clean_text(values[1]) or "B2C").upper(),
            "order_date": values[4],
            "order_name": order_name,
            "invoice_name": _clean_text(values[6]),
            "total_qty": _float_or_zero(values[7]),
            "amount_total": _money(values[8]),
            "discount": _money(values[9]),
            "net_total": _money(values[10]),
            "payment_date": values[11] or values[4],
            "cash_received": _money(values[12]),
            "ibft_received": _money(values[13]),
            "balance_due": _money(values[14]),
            "payment_status": _clean_text(values[15]) or "Paid",
            "state": _clean_text(values[16]) or "Delivered to Customer",
            "source_file": str(xlsx_path),
        }
        for key in totals:
            totals[key] += row[key]

        bucket = grouped_orders.setdefault(
            order_name,
            {
                "order_name": order_name,
                "branch": row["branch"],
                "customer_type": row["customer_type"],
                "order_date": row["order_date"],
                "state": row["state"],
                "source_file": row["source_file"],
                "row_indexes": [],
                "lines": [],
                "amount_total": 0.0,
                "discount": 0.0,
                "net_total": 0.0,
                "cash_received": 0.0,
                "ibft_received": 0.0,
                "balance_due": 0.0,
            },
        )
        bucket["row_indexes"].append(row_index)
        bucket["lines"].append(row)
        for key in ("amount_total", "discount", "net_total", "cash_received", "ibft_received", "balance_due"):
            bucket[key] += row[key]
        if row["order_date"] and bucket["order_date"] and row["order_date"] < bucket["order_date"]:
            bucket["order_date"] = row["order_date"]

    for bucket in grouped_orders.values():
        for key in ("amount_total", "discount", "net_total", "cash_received", "ibft_received", "balance_due"):
            bucket[key] = _money(bucket[key])

    return {
        "headers": headers[:17],
        "rows": list(grouped_orders.values()),
        "totals": {key: _money(value) for key, value in totals.items()},
    }


def _ensure_record(model, domain, values):
    record = model.search(domain, limit=1)
    if record:
        return record
    return model.create(values)


def _prepare_env(env):
    company = env["res.company"].sudo().search([("name", "=", COMPANY_NAME)], limit=1)
    if not company:
        raise UserError("Missing company for legacy import: %s" % COMPANY_NAME)

    ctx = dict(env.context, allowed_company_ids=[company.id], force_company=company.id)
    company_env = api.Environment(env.cr, env.uid, ctx)

    Partner = company_env["res.partner"].sudo()
    Category = company_env["aimaze.laundry.service.category"].sudo()
    Service = company_env["aimaze.laundry.service"].sudo()
    Colour = company_env["goldverse.laundry.colour"].sudo()
    Topup = company_env["goldverse.laundry.topup"].sudo()
    Branch = company_env["aimaze.laundry.branch"].sudo()
    Journal = company_env["account.journal"].sudo()

    partner = Partner.with_context(goldverse_skip_duplicate_mobile_check=True).search(
        [("name", "=", LEGACY_CUSTOMER_NAME)],
        limit=1,
    )
    if not partner:
        partner = Partner.with_context(goldverse_skip_duplicate_mobile_check=True).create(
            {
                "name": LEGACY_CUSTOMER_NAME,
                "mobile": LEGACY_CUSTOMER_MOBILE,
                "phone": LEGACY_CUSTOMER_MOBILE,
                "customer_rank": 1,
                "company_id": company.id,
                "goldverse_customer_category": "b2c",
                "laundry_customer_type": "b2c",
            }
        )

    category = _ensure_record(Category, [("name", "=", LEGACY_CATEGORY_NAME)], {"name": LEGACY_CATEGORY_NAME})
    colour = (
        Colour.search([("name", "ilike", "multi")], limit=1)
        or Colour.search([], limit=1)
        or _ensure_record(Colour, [("name", "=", "Legacy Unknown")], {"name": "Legacy Unknown"})
    )
    topup = (
        Topup.search([("name", "=", "Folded Pack")], limit=1)
        or Topup.search([], limit=1)
    )
    if not topup:
        raise UserError("Missing Add On records for legacy import.")

    service = Service.search([("name", "=", LEGACY_SERVICE_NAME)], limit=1)
    if not service:
        service = Service.create(
            {
                "name": LEGACY_SERVICE_NAME,
                "category_id": category.id,
                "list_price": 0.0,
                "goldverse_base_price": 0.0,
            }
        )

    branch = (
        Branch.search([("name", "=", "EME Branch")], limit=1)
        or Branch.search([("name", "ilike", "EME")], limit=1)
        or Branch.search([], limit=1)
    )
    if not branch:
        raise UserError("Missing EME branch for legacy import.")

    sale_journal = Journal.search([("company_id", "=", company.id), ("type", "=", "sale")], limit=1)
    cash_journal = Journal.search([("company_id", "=", company.id), ("name", "=", "Cash")], limit=1)
    ibft_journal = Journal.search([("company_id", "=", company.id), ("name", "=", "IBFT")], limit=1)
    if not sale_journal:
        raise UserError("Missing Sales journal for legacy import.")
    if not cash_journal or not cash_journal.inbound_payment_method_line_ids:
        raise UserError("Missing Cash journal or inbound payment method line for legacy import.")
    if not ibft_journal or not ibft_journal.inbound_payment_method_line_ids:
        raise UserError("Missing IBFT journal or inbound payment method line for legacy import.")

    return {
        "env": company_env,
        "company": company,
        "partner": partner,
        "category": category,
        "service": service,
        "colour": colour,
        "topup": topup,
        "branch": branch,
        "sale_journal": sale_journal,
        "cash_journal": cash_journal,
        "cash_method_line": cash_journal.inbound_payment_method_line_ids[:1],
        "ibft_journal": ibft_journal,
        "ibft_method_line": ibft_journal.inbound_payment_method_line_ids[:1],
    }


def _validate_rows(rows):
    problems = []
    for row in rows:
        payment_total = _money(row["cash_received"] + row["ibft_received"])
        if abs(payment_total - row["net_total"]) > 0.01:
            problems.append(
                "Order %(order_name)s payment total %(payment_total).2f != net %(net_total).2f"
                % {
                    "order_name": row["order_name"],
                    "payment_total": payment_total,
                    "net_total": row["net_total"],
                }
            )
        if row["balance_due"] and abs(row["balance_due"]) > 0.01:
            problems.append("Order %(order_name)s has non-zero balance due %(balance_due).2f" % row)
        if row["customer_type"] != "B2C":
            problems.append("Order %(order_name)s customer type is %(customer_type)s" % row)
    if problems:
        raise ValidationError("\n".join(problems[:100]))


def _reconcile_payment_to_invoice(invoice, payment):
    invoice_lines = invoice.line_ids.filtered(
        lambda line: line.account_id.account_type == "asset_receivable" and not line.reconciled
    )
    payment_lines = payment.move_id.line_ids.filtered(
        lambda line: line.account_id.account_type == "asset_receivable" and not line.reconciled
    )
    (invoice_lines + payment_lines).reconcile()


def _existing_order_names(order_model):
    return set(order_model.search([]).mapped("name"))


def _create_payment(prepared, order, invoice, amount, payment_date, journal, method_line, memo):
    if amount <= 0:
        return False
    receivable_line = invoice.line_ids.filtered(
        lambda line: line.account_id.account_type == "asset_receivable"
    )[:1]
    payment = prepared["env"]["account.payment"].sudo().create(
        {
            "company_id": prepared["company"].id,
            "payment_type": "inbound",
            "partner_type": "customer",
            "partner_id": prepared["partner"].id,
            "aimaze_laundry_order_id": order.id,
            "amount": amount,
            "date": payment_date,
            "journal_id": journal.id,
            "payment_method_line_id": method_line.id,
            "memo": memo,
            "destination_account_id": receivable_line.account_id.id if receivable_line else False,
        }
    )
    payment.action_post()
    _reconcile_payment_to_invoice(invoice, payment)
    return payment


def _patch_order_report_fields(order, row):
    latest_payment_date = False
    for line in row["lines"]:
        payment_date = _date_or_false(line["payment_date"])
        if payment_date and (not latest_payment_date or payment_date > latest_payment_date):
            latest_payment_date = payment_date
    values = {
        "goldverse_report_base_price": row["amount_total"],
        "goldverse_report_discount": row["discount"],
        "goldverse_report_net_price": row["net_total"],
        "goldverse_report_priority_charges": 0.0,
        "goldverse_report_total_excl_tax": row["net_total"],
        "goldverse_report_payment_date": latest_payment_date,
        "goldverse_report_cash_received": row["cash_received"],
        "goldverse_report_ibft_received": row["ibft_received"],
        "delivery_charge": 0.0,
    }
    order.with_context(
        goldverse_allow_locked_order_write=True,
        goldverse_skip_required_validation=True,
    ).write(values)


def _patch_order_line_breakdown(order, row):
    ordered_lines = order.line_ids.sorted("id")
    for line, line_row in zip(ordered_lines, row["lines"]):
        line.with_context(
            goldverse_allow_locked_order_write=True,
            goldverse_refreshing_amounts=True,
        ).write(
            {
                "goldverse_base_price": line_row["amount_total"],
                "goldverse_discount_amount": line_row["discount"],
                "goldverse_net_price": line_row["net_total"],
                "goldverse_priority_charge": 0.0,
            }
        )


def _import_rows(env, workbook_path, commit=False, limit=0):
    payload = _load_rows(workbook_path)
    rows = payload["rows"]
    _validate_rows(rows)
    prepared = _prepare_env(env)
    Order = prepared["env"]["aimaze.laundry.order"].sudo()
    existing_order_names = _existing_order_names(Order)
    chunk_size = int(os.environ.get("GOLDVERSE_LEGACY_B2C_CHUNK_SIZE") or 200)
    selected_rows = rows[:limit or None]

    created_orders = 0
    created_invoices = 0
    created_payments = 0
    skipped = 0
    imported_amounts = {
        "amount_total": 0.0,
        "discount": 0.0,
        "net_total": 0.0,
        "cash_received": 0.0,
        "ibft_received": 0.0,
    }
    first_invoice_name = False
    last_invoice_name = False
    validation_only = not commit

    for chunk_start in range(0, len(selected_rows), chunk_size):
        chunk_rows = selected_rows[chunk_start : chunk_start + chunk_size]
        for row in chunk_rows:
            if row["order_name"] in existing_order_names:
                skipped += 1
                continue

            order_date_dt = _datetime_or_false(row["order_date"]) or fields.Datetime.now()
            order_date_str = fields.Datetime.to_string(order_date_dt)
            delivery_datetime_str = _legacy_delivery_datetime(row["order_date"]) or order_date_str
            line_commands = []
            for line_index, line_row in enumerate(row["lines"], start=1):
                line_commands.append(
                    (
                        0,
                        0,
                        {
                            "service_id": prepared["service"].id,
                            "quantity": 1,
                            "goldverse_priority": "normal",
                            "goldverse_category_id": prepared["category"].id,
                            "goldverse_colour_ids": [(6, 0, [prepared["colour"].id])],
                            "goldverse_topup_ids": [(6, 0, [prepared["topup"].id])],
                            "name": "%s #%s" % (LEGACY_SERVICE_NAME, line_index),
                        },
                    )
                )

            order = Order.create(
                {
                    "name": row["order_name"],
                    "barcode": row["order_name"],
                    "partner_id": prepared["partner"].id,
                    "mobile_partner_id": prepared["partner"].id,
                    "mobile": prepared["partner"].mobile,
                    "email": False,
                    "customer_type": "b2c",
                    "source": DEFAULT_SOURCE,
                    "priority": "normal",
                    "branch_id": prepared["branch"].id,
                    "order_date": order_date_str,
                    "expected_delivery_datetime": delivery_datetime_str,
                    "line_ids": line_commands,
                }
            )

            for line, line_row in zip(order.line_ids.sorted("id"), row["lines"]):
                line.with_context(goldverse_refreshing_amounts=True).write(
                    {
                        "quantity": 1,
                        "unit_price": line_row["net_total"],
                        "discount": 0.0,
                        "tax_ids": [(5, 0, 0)],
                    }
                )

            order.action_create_order()
            order.action_create_invoice()
            invoice = order.invoice_id
            invoice.write(
                {
                    "invoice_date": _date_or_false(row["order_date"]),
                    "ref": row["order_name"],
                    "invoice_origin": row["order_name"],
                    "payment_reference": row["order_name"],
                    "narration": "%s\nSource file rows: %s\nWorkbook: %s"
                    % (IMPORT_MARKER, ", ".join(str(item) for item in row["row_indexes"]), row["source_file"]),
                }
            )
            if invoice.state == "draft":
                invoice.action_post()
            created_invoices += 1
            if not first_invoice_name:
                first_invoice_name = invoice.name
            last_invoice_name = invoice.name

            for line_row in row["lines"]:
                payment_date = _date_or_false(line_row["payment_date"]) or _date_or_false(line_row["order_date"])
                cash_payment = _create_payment(
                    prepared,
                    order,
                    invoice,
                    line_row["cash_received"],
                    payment_date,
                    prepared["cash_journal"],
                    prepared["cash_method_line"],
                    "%s CASH ROW %s" % (row["order_name"], line_row["row_index"]),
                )
                if cash_payment:
                    created_payments += 1
                ibft_payment = _create_payment(
                    prepared,
                    order,
                    invoice,
                    line_row["ibft_received"],
                    payment_date,
                    prepared["ibft_journal"],
                    prepared["ibft_method_line"],
                    "%s IBFT ROW %s" % (row["order_name"], line_row["row_index"]),
                )
                if ibft_payment:
                    created_payments += 1

            latest_payment_value = row["order_date"]
            for line_row in row["lines"]:
                if line_row["payment_date"] and (not latest_payment_value or line_row["payment_date"] > latest_payment_value):
                    latest_payment_value = line_row["payment_date"]
            delivered_datetime = _payment_datetime(latest_payment_value or row["order_date"])
            order.with_context(
                goldverse_allow_locked_order_write=True,
                goldverse_skip_required_validation=True,
            ).write(
                {
                    "warehouse_collected_datetime": order_date_str,
                    "warehouse_received_datetime": order_date_str,
                    "goldverse_delivered_to_customer": True,
                    "goldverse_actual_delivery_datetime": delivered_datetime,
                }
            )
            order.with_context(
                goldverse_allow_locked_order_write=True,
                goldverse_skip_required_validation=True,
            )._set_state("paid")
            _patch_order_line_breakdown(order, row)
            _patch_order_report_fields(order, row)

            created_orders += 1
            existing_order_names.add(row["order_name"])
            for key in imported_amounts:
                imported_amounts[key] += row[key]

        if commit:
            env.cr.commit()
        print(
            json.dumps(
                {
                    "chunk_start": chunk_start + 1,
                    "chunk_end": min(chunk_start + len(chunk_rows), len(selected_rows)),
                    "created_orders": created_orders,
                    "created_invoices": created_invoices,
                    "created_payments": created_payments,
                    "skipped_existing_orders": skipped,
                },
                sort_keys=True,
            ),
            flush=True,
        )

    result = {
        "commit": bool(commit),
        "validation_only": validation_only,
        "created_orders": created_orders,
        "created_invoices": created_invoices,
        "created_payments": created_payments,
        "skipped_existing_orders": skipped,
        "workbook_rows": len(selected_rows),
        "workbook_totals": payload["totals"],
        "imported_totals": {key: _money(value) for key, value in imported_amounts.items()},
        "first_invoice_name": first_invoice_name,
        "last_invoice_name": last_invoice_name,
    }
    if commit:
        env.cr.commit()
        result["transaction"] = "committed"
    else:
        env.cr.rollback()
        result["transaction"] = "rolled_back"
    return result


workbook_env = os.environ.get("GOLDVERSE_LEGACY_B2C_XLSX")
if not workbook_env:
    raise UserError("Set GOLDVERSE_LEGACY_B2C_XLSX before running this script.")

workbook_path = Path(workbook_env)
if not workbook_path.exists():
    raise UserError("Workbook not found: %s" % workbook_path)

limit = int(os.environ.get("GOLDVERSE_LEGACY_B2C_LIMIT") or 0)
commit = os.environ.get("GOLDVERSE_LEGACY_B2C_COMMIT") == "1"
result = _import_rows(env, workbook_path, commit=commit, limit=limit)
print(json.dumps(result, indent=2, sort_keys=True, default=str))
